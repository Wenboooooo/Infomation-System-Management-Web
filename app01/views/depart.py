from django.shortcuts import render, redirect, HttpResponse
from app01 import models
from app01.utils.pagination import Pagination


# Create your views here.
def depart_list(request):
    """Department List"""
    # Acquire department list from the database
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset, page_size=5)
    context = {"queryset": page_object.page_queryset,
               "page_string": page_object.html(), }

    return render(request, "depart_list.html", context)


def depart_add(request):
    """Add Department"""
    if request.method == "GET":
        return render(request, "depart_add.html")
    # Acquire info submitted by post method
    title = request.POST.get("title")
    # Save it to the database
    models.Department.objects.create(title=title)
    # Redirect back to the department list page
    return redirect("/depart/list/")


def depart_delete(request):
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """modify department"""
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"old_title": row_object.title})
    # Acquire the new title submitted by users
    new_title = request.POST.get("title")
    # Update data
    models.Department.objects.filter(id=nid).update(title=new_title)
    return redirect("/depart/list/")


def depart_index(request):
    return render(request, "index.html")



def depart_multi(request):
    """Upload Files (excel)"""

    from django.core.files.uploadedfile import InMemoryUploadedFile
    from openpyxl import load_workbook

    # Get the object uploaded by the user
    file_object = request.FILES.get("exc")
    print(type(file_object))

    # Deliver it to load_workbook package to read the content of the object
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # Get data line by line
    for row in sheet.iter_rows(min_row=1):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if exists:
            continue
        models.Department.objects.create(title=text)

    return redirect('/depart/list/')











